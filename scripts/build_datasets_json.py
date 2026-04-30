"""
Convert the INCLUDE_Datasets sheet of the master inventory XLSX into a clean
JSON array at src/data/datasets.json.
"""
from __future__ import annotations

import json
import re
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parent.parent
XLSX = REPO / "src" / "data" / "stroke_scrna_master_inventory.xlsx"
OUT = REPO / "src" / "data" / "datasets.json"

FIELDS = [
    "accession", "first_author", "year", "journal", "pmid", "doi",
    "geo_url", "paper_url", "organism", "assay", "sex", "age",
    "stroke_model", "model_description", "timepoints", "genetic_background",
    "region_dissected", "control_type", "n_cells_atlas", "data_format",
    "download_status", "notes", "stroke_model_verified",
    "reperfusion_time_minutes", "tissue_region", "tissue_region_detail",
    "reperfusion_verification",
]
INT_FIELDS = {"year", "pmid"}


def parse_n_cells(value):
    """Extract the first integer from cell-count strings like
    '60,202 (WT only; full = 155,804)'. Returns int or None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    m = re.search(r"[\d,]+", s)
    if not m:
        return None
    return int(m.group(0).replace(",", ""))


def parse_reperfusion(value):
    """Reperfusion time in minutes — int or None."""
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s or s.lower() in {"na", "n/a", "none", "null"}:
        return None
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else None


def parse_int(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    m = re.search(r"\d+", s)
    return int(m.group(0)) if m else None


def normalize_string(value):
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["INCLUDE_Datasets"]

    headers = [c.value for c in ws[1]]
    col_index = {h: i for i, h in enumerate(headers) if h}

    missing = [f for f in FIELDS if f not in col_index]
    if missing:
        raise SystemExit(f"Missing columns in sheet: {missing}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        record = {}
        for field in FIELDS:
            raw = row[col_index[field]]
            if field == "n_cells_atlas":
                record[field] = parse_n_cells(raw)
            elif field == "reperfusion_time_minutes":
                record[field] = parse_reperfusion(raw)
            elif field in INT_FIELDS:
                record[field] = parse_int(raw)
            else:
                record[field] = normalize_string(raw)
        rows.append(record)

    OUT.write_text(json.dumps(rows, indent=2, ensure_ascii=False) + "\n")

    print(f"Wrote {len(rows)} records to {OUT.relative_to(REPO)}")
    print("\nFirst two entries:")
    print(json.dumps(rows[:2], indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
